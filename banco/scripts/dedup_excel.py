"""
dedup_excel.py — identifica questões do Excel sem correspondente em banco/questoes/q-*.md

Saída: banco/meta/EXCEL_UNICOS.json
"""

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "banco_provas.xlsx"
QUESTOES_DIR = ROOT / "questoes"
OUTPUT_PATH = ROOT / "meta" / "EXCEL_UNICOS.json"

MATCH_PREFIX_LEN = 80


def normalize(text: str) -> str:
    if not text:
        return ""
    text = str(text)
    # remove acentos
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text)   # remove pontuação
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_enunciado_from_md(path: Path) -> str:
    """Extrai o corpo do enunciado: texto após '## Enunciado' até a próxima seção ##."""
    content = path.read_text(encoding="utf-8", errors="replace")
    # pular frontmatter YAML
    if content.startswith("---"):
        end = content.find("---", 3)
        if end != -1:
            content = content[end + 3:]
    match = re.search(r"##\s*Enunciado\s*\n(.*?)(?=\n##|\Z)", content, re.DOTALL)
    if match:
        return match.group(1).strip()
    # fallback: primeiro parágrafo não-vazio fora do frontmatter
    for line in content.splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            return line
    return ""


def build_md_index() -> dict[str, str]:
    """Retorna dict prefix80 -> arquivo (para detecção de duplicatas)."""
    index = {}
    for path in sorted(QUESTOES_DIR.glob("q-*.md")):
        enunciado = extract_enunciado_from_md(path)
        key = normalize(enunciado)[:MATCH_PREFIX_LEN]
        if key:
            index[key] = path.name
    return index


def read_excel_rows() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        if not data.get("Enunciado"):
            continue
        rows.append({"_row": i, **data})
    return rows


def main():
    print(f"Lendo {EXCEL_PATH.name}...")
    excel_rows = read_excel_rows()
    print(f"  {len(excel_rows)} linhas com enunciado no Excel")

    print(f"Indexando q-*.md em {QUESTOES_DIR}...")
    md_index = build_md_index()
    print(f"  {len(md_index)} arquivos .md indexados")

    unicos = []
    matched = 0

    for row in excel_rows:
        enunciado_raw = str(row.get("Enunciado", ""))
        key = normalize(enunciado_raw)[:MATCH_PREFIX_LEN]
        if key in md_index:
            matched += 1
            continue

        alternativas = {
            "A": row.get("A") or "",
            "B": row.get("B") or "",
            "C": row.get("C") or "",
            "D": row.get("D") or "",
            "E": row.get("E") or "",
        }

        unicos.append({
            "id_excel": row.get("ID"),
            "enunciado": enunciado_raw,
            "alternativas": alternativas,
            "gabarito_claude": row.get("Gabarito_Claude") or "",
            "justificativa": row.get("Justificativa") or "",
            "tipo": row.get("Tipo") or "",
            "tema": row.get("Tema") or "",
            "uc": row.get("UC") or "",
            "dificuldade": row.get("Dificuldade") or "",
            "tem_imagem": bool(row.get("Tem_Imagem")),
            "imagem": row.get("Imagem") or "",
        })

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(unicos, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nResultado:")
    print(f"  Com correspondente em .md : {matched}")
    print(f"  Únicas (sem correspondente): {len(unicos)}")
    print(f"  Salvo em: {OUTPUT_PATH}")
    return unicos


if __name__ == "__main__":
    main()
