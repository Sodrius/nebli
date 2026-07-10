"""
banco/ingest.py — pipeline de ingestao de questoes para o banco_provas.xlsx.

Uso:
    python banco/ingest.py < payload.json
    ou
    cat payload.json | python banco/ingest.py

Payload JSON esperado:
{
  "pdf": "PROVA_UC1_xxx.pdf",           # nome do PDF (sem path)
  "uc": "UC1" | "UC2" | "SEM-UC",
  "pdf_path": "_material/provas/UC1/PROVA_UC1_xxx.pdf",
  "questoes": [
    {
      "pagina": 3,
      "tipo": "MC" | "CE" | "Discursiva",
      "tema": "Bioquimica - Cadeia respiratoria",
      "enunciado": "...",
      "A": "...", "B": "...", "C": "...", "D": "...", "E": "...",
      "gabarito": "C",
      "justificativa": "Mecanismo: ...",
      "dificuldade": "Facil" | "Intermediario" | "Dificil",
      "tem_imagem": true | false
    }, ...
  ]
}

Responsabilidades:
- dedup por hash do enunciado normalizado (lower, sem acento, sem pontuacao, 150 chars)
- gera PNG 180 dpi para questoes com imagem (reusa se ja existir)
- mantem progresso.json (lista de PDFs processados + set de hashes vistos)
- cria/abre banco_provas.xlsx, aplica estilo no header, freeze pane, wrap
- salva a cada chamada (1 PDF por chamada => salva por PDF)
"""

import sys
import os
import json
import hashlib
import unicodedata
import re
import datetime
from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
BANCO = ROOT / "banco"
XLSX = BANCO / "banco_provas.xlsx"
PROGRESSO = BANCO / "progresso.json"
FIGURAS = BANCO / "figuras"
FIGURAS.mkdir(exist_ok=True)

HEADER = [
    "ID", "UC", "Tema", "Fonte", "Pagina_Origem", "Tipo", "Enunciado",
    "A", "B", "C", "D", "E", "Gabarito_Claude", "Justificativa",
    "Dificuldade", "Tem_Imagem", "Imagem", "Status", "Data_Processamento",
]

COL_WIDTHS = {
    "A": 6, "B": 10, "C": 36, "D": 36, "E": 8, "F": 8, "G": 70,
    "H": 28, "I": 28, "J": 28, "K": 28, "L": 28, "M": 10, "N": 70,
    "O": 14, "P": 8, "Q": 32, "R": 12, "S": 18,
}
WRAP_COLS = {"G", "H", "I", "J", "K", "L", "N", "Q"}


def normalize(text: str) -> str:
    t = text.lower().strip()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"[^a-z0-9 ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t[:300]


def hash_enunciado(text: str) -> str:
    return hashlib.sha256(normalize(text).encode("utf-8")).hexdigest()[:16]


def load_progresso():
    if PROGRESSO.exists():
        with open(PROGRESSO, "r", encoding="utf-8") as f:
            data = json.load(f)
        data.setdefault("pdfs_processados", [])
        data.setdefault("hashes", {})
        data.setdefault("ultimo_id", 0)
        return data
    return {"pdfs_processados": [], "hashes": {}, "ultimo_id": 0}


def save_progresso(prog):
    with open(PROGRESSO, "w", encoding="utf-8") as f:
        json.dump(prog, f, ensure_ascii=False, indent=2)


def open_or_create_xlsx():
    if XLSX.exists():
        wb = load_workbook(XLSX)
        ws = wb["Questoes"]
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Questoes"
    ws.append(HEADER)
    fill = PatternFill("solid", fgColor="1B2A4A")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = align
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(XLSX)
    return wb, ws


def extract_page_png(pdf_path: str, page_num: int, out_path: Path):
    if out_path.exists():
        return
    doc = fitz.open(pdf_path)
    page = doc[page_num - 1]
    matrix = fitz.Matrix(180 / 72, 180 / 72)
    pix = page.get_pixmap(matrix=matrix)
    pix.save(str(out_path))
    doc.close()


def main():
    payload = json.load(sys.stdin)
    pdf_name = payload["pdf"]
    uc = payload["uc"]
    pdf_path = payload["pdf_path"]
    questoes = payload["questoes"]

    prog = load_progresso()
    wb, ws = open_or_create_xlsx()

    pdf_stem = Path(pdf_name).stem
    novos = 0
    duplicatas = 0
    com_imagem = 0
    hashes_pdf = []

    next_id = prog["ultimo_id"] + 1
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    align_wrap = Alignment(wrap_text=True, vertical="top")
    align_normal = Alignment(vertical="top")

    for q in questoes:
        h = hash_enunciado(q["enunciado"])
        if h in prog["hashes"]:
            duplicatas += 1
            origem = prog["hashes"][h]
            print(f"  DUPLICATA p.{q['pagina']:>3} hash={h} (origem: {origem})")
            continue

        imagem_path = ""
        if q.get("tem_imagem"):
            png_name = f"{pdf_stem}_p{q['pagina']:02d}.png"
            png_path = FIGURAS / png_name
            try:
                extract_page_png(pdf_path, q["pagina"], png_path)
                imagem_path = f"figuras/{png_name}"
                com_imagem += 1
            except Exception as e:
                print(f"  ! erro PNG p.{q['pagina']}: {e}")

        row = [
            next_id,
            uc,
            q.get("tema", ""),
            pdf_name,
            q["pagina"],
            q["tipo"],
            q["enunciado"],
            q.get("A", ""), q.get("B", ""), q.get("C", ""),
            q.get("D", ""), q.get("E", ""),
            q.get("gabarito", ""),
            q.get("justificativa", ""),
            q.get("dificuldade", ""),
            "TRUE" if q.get("tem_imagem") else "FALSE",
            imagem_path,
            "Pendente",
            now,
        ]
        ws.append(row)
        r = ws.max_row
        for col_letter in COL_WIDTHS.keys():
            cell = ws[f"{col_letter}{r}"]
            cell.alignment = align_wrap if col_letter in WRAP_COLS else align_normal

        prog["hashes"][h] = pdf_name
        hashes_pdf.append(h)
        next_id += 1
        novos += 1

    prog["ultimo_id"] = next_id - 1
    if pdf_name not in prog["pdfs_processados"]:
        prog["pdfs_processados"].append(pdf_name)

    wb.save(XLSX)
    save_progresso(prog)

    print(f"OK {pdf_name}: novas={novos}, duplicatas={duplicatas}, com_imagem={com_imagem}, total_banco={prog['ultimo_id']}")


if __name__ == "__main__":
    main()
